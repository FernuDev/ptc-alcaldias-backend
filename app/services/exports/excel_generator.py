"""Excel export with two sheets: data + mini dashboard.

Sheet 1 "Datos": Full data table with headers and auto-filters.
Sheet 2 "Dashboard": KPIs summary, distribution tables, and charts.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GUINDA = "9F2241"
VERDE = "2D7A4F"
DORADO = "BC955C"
DARK = "1E293B"
LIGHT = "F1F5F9"
WHITE = "FFFFFF"

_header_font = Font(bold=True, color=WHITE, size=10)
_header_fill = PatternFill(start_color=GUINDA, end_color=GUINDA, fill_type="solid")
_header_border = Border(bottom=Side(style="thin", color="94A3B8"))
_title_font = Font(bold=True, color=DARK, size=14)
_kpi_value_font = Font(bold=True, color=GUINDA, size=20)
_kpi_label_font = Font(color="64748B", size=9)
_alt_fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")


def _style_header_row(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.border = _header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24


def _style_data_rows(ws, nrows: int, ncols: int):
    for row in range(2, nrows + 2):
        if row % 2 == 0:
            for col in range(1, ncols + 1):
                ws.cell(row=row, column=col).fill = _alt_fill
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)


def _auto_width(ws, ncols: int, max_width: int = 35):
    for col in range(1, ncols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3


# ═══════════════════════════════════════════════════════════════════════════
# Reportes Excel
# ═══════════════════════════════════════════════════════════════════════════

def generate_reportes_excel(
    *,
    reportes: list[dict],
    kpis: dict,
    dist_categoria: list[dict],
    dist_estado: list[dict],
    top_colonias: list[dict],
    tenant_name: str,
) -> bytes:
    wb = Workbook()

    # ── Sheet 1: Datos ──
    ws = wb.active
    ws.title = "Datos"
    headers = [
        "Folio", "Categoría", "Estado", "Prioridad", "Fuente",
        "Colonia", "Título", "Ciudadano", "Cuadrilla",
        "Fecha creación", "Fecha cierre", "Horas atención",
        "Costo estimado", "Gasto real", "Lng", "Lat",
    ]
    ws.append(headers)
    for r in reportes:
        ws.append([
            r.get("folio", ""),
            r.get("categoria_id", ""),
            r.get("estado", ""),
            r.get("prioridad", ""),
            r.get("fuente", ""),
            r.get("colonia_nombre", ""),
            r.get("titulo", ""),
            r.get("ciudadano_nombre", ""),
            r.get("cuadrilla_id", ""),
            str(r.get("fecha_creacion", ""))[:19],
            str(r.get("fecha_cierre", "") or "")[:19],
            float(r["tiempo_atencion_horas"]) if r.get("tiempo_atencion_horas") else None,
            float(r["costo_estimado"]) if r.get("costo_estimado") else None,
            float(r["gasto_real"]) if r.get("gasto_real") else None,
            r.get("lng"),
            r.get("lat"),
        ])
    _style_header_row(ws, len(headers))
    _style_data_rows(ws, len(reportes), len(headers))
    _auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(reportes) + 1}"
    ws.freeze_panes = "A2"

    # ── Sheet 2: Dashboard ──
    ds = wb.create_sheet("Dashboard")

    # Title
    ds.merge_cells("A1:F1")
    ds["A1"] = f"Dashboard · {tenant_name}"
    ds["A1"].font = _title_font
    ds.row_dimensions[1].height = 30

    # KPIs row
    kpi_items = [
        ("Activos", kpis.get("activos", 0)),
        ("Resueltos", kpis.get("resueltos", 0)),
        ("Tiempo prom. (días)", f'{kpis.get("tiempo_promedio_dias", 0):.1f}'),
        ("En riesgo SLA", kpis.get("en_riesgo_sla", 0)),
        ("Total", kpis.get("total_rango", 0)),
        ("% Resolución", f'{kpis.get("pct_resueltos", 0):.1f}%'),
    ]
    for i, (label, value) in enumerate(kpi_items):
        col = i + 1
        ds.cell(row=3, column=col, value=value).font = _kpi_value_font
        ds.cell(row=3, column=col).alignment = Alignment(horizontal="center")
        ds.cell(row=4, column=col, value=label).font = _kpi_label_font
        ds.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        ds.column_dimensions[get_column_letter(col)].width = 18

    # Distribution by category table + chart
    ds["A7"] = "Distribución por categoría"
    ds["A7"].font = Font(bold=True, color=DARK, size=11)
    ds["A8"] = "Categoría"
    ds["B8"] = "Cantidad"
    ds["C8"] = "%"
    for cell in [ds["A8"], ds["B8"], ds["C8"]]:
        cell.font = _header_font
        cell.fill = _header_fill
    for i, d in enumerate(dist_categoria):
        row = 9 + i
        ds.cell(row=row, column=1, value=d["label"])
        ds.cell(row=row, column=2, value=d["count"])
        ds.cell(row=row, column=3, value=d["pct"])

    if dist_categoria:
        cat_chart = PieChart()
        cat_chart.title = "Por categoría"
        cat_chart.width = 14
        cat_chart.height = 10
        end_row = 8 + len(dist_categoria)
        cat_chart.add_data(Reference(ds, min_col=2, min_row=8, max_row=end_row), titles_from_data=True)
        cat_chart.set_categories(Reference(ds, min_col=1, min_row=9, max_row=end_row))
        cat_chart.dataLabels = DataLabelList()
        cat_chart.dataLabels.showPercent = True
        cat_chart.dataLabels.showVal = False
        ds.add_chart(cat_chart, "E7")

    # Distribution by state table + chart
    state_start = 9 + len(dist_categoria) + 2
    ds.cell(row=state_start, column=1, value="Distribución por estado").font = Font(bold=True, color=DARK, size=11)
    ds.cell(row=state_start + 1, column=1, value="Estado").font = _header_font
    ds.cell(row=state_start + 1, column=1).fill = _header_fill
    ds.cell(row=state_start + 1, column=2, value="Cantidad").font = _header_font
    ds.cell(row=state_start + 1, column=2).fill = _header_fill
    for i, d in enumerate(dist_estado):
        row = state_start + 2 + i
        ds.cell(row=row, column=1, value=d["label"])
        ds.cell(row=row, column=2, value=d["count"])

    if dist_estado:
        est_chart = BarChart()
        est_chart.title = "Por estado"
        est_chart.type = "col"
        est_chart.width = 14
        est_chart.height = 10
        end_row = state_start + 1 + len(dist_estado)
        est_chart.add_data(Reference(ds, min_col=2, min_row=state_start + 1, max_row=end_row), titles_from_data=True)
        est_chart.set_categories(Reference(ds, min_col=1, min_row=state_start + 2, max_row=end_row))
        ds.add_chart(est_chart, f"E{state_start}")

    # Top colonias
    col_start = state_start + len(dist_estado) + 4
    ds.cell(row=col_start, column=1, value="Top colonias").font = Font(bold=True, color=DARK, size=11)
    ds.cell(row=col_start + 1, column=1, value="Colonia").font = _header_font
    ds.cell(row=col_start + 1, column=1).fill = _header_fill
    ds.cell(row=col_start + 1, column=2, value="Reportes").font = _header_font
    ds.cell(row=col_start + 1, column=2).fill = _header_fill
    for i, c in enumerate(top_colonias):
        row = col_start + 2 + i
        ds.cell(row=row, column=1, value=c["colonia_nombre"])
        ds.cell(row=row, column=2, value=c["count"])

    if top_colonias:
        col_chart = BarChart()
        col_chart.title = "Top colonias"
        col_chart.type = "bar"
        col_chart.width = 14
        col_chart.height = 10
        end_row = col_start + 1 + len(top_colonias)
        col_chart.add_data(Reference(ds, min_col=2, min_row=col_start + 1, max_row=end_row), titles_from_data=True)
        col_chart.set_categories(Reference(ds, min_col=1, min_row=col_start + 2, max_row=end_row))
        ds.add_chart(col_chart, f"E{col_start}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# Obras Excel
# ═══════════════════════════════════════════════════════════════════════════

def generate_obras_excel(*, obras: list[dict], tenant_name: str) -> bytes:
    wb = Workbook()

    # Sheet 1: Datos
    ws = wb.active
    ws.title = "Datos"
    headers = [
        "Folio", "Nombre", "Categoría", "Estado", "Prioridad",
        "Colonia", "Contratista", "Avance %",
        "Presupuesto autorizado", "Presupuesto ejercido",
        "Fecha inicio", "Fecha fin estimada", "Fecha fin real",
        "Responsable",
    ]
    ws.append(headers)
    for o in obras:
        ws.append([
            o.get("folio", ""),
            o.get("nombre", ""),
            o.get("categoria_id", ""),
            o.get("estado", ""),
            o.get("prioridad", ""),
            o.get("colonia_nombre", ""),
            o.get("contratista_id", ""),
            o.get("avance_pct", 0),
            float(o["presupuesto_autorizado"]) if o.get("presupuesto_autorizado") else None,
            float(o["presupuesto_ejercido"]) if o.get("presupuesto_ejercido") else None,
            str(o.get("fecha_inicio", ""))[:10],
            str(o.get("fecha_fin_estimada", ""))[:10],
            str(o.get("fecha_fin_real", "") or "")[:10],
            o.get("responsable_nombre", ""),
        ])
    _style_header_row(ws, len(headers))
    _style_data_rows(ws, len(obras), len(headers))
    _auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(obras) + 1}"
    ws.freeze_panes = "A2"

    # Sheet 2: Dashboard
    ds = wb.create_sheet("Dashboard")
    ds.merge_cells("A1:E1")
    ds["A1"] = f"Obras · {tenant_name}"
    ds["A1"].font = _title_font

    total = len(obras)
    activas = sum(1 for o in obras if o.get("estado") not in ("concluida", "suspendida"))
    pres_total = sum(float(o.get("presupuesto_autorizado") or 0) for o in obras)
    pres_ejercido = sum(float(o.get("presupuesto_ejercido") or 0) for o in obras)
    avg_avance = sum(o.get("avance_pct", 0) for o in obras) / max(total, 1)

    kpis = [
        ("Total obras", total), ("Activas", activas),
        ("Pres. autorizado", f"${pres_total:,.0f}"),
        ("Pres. ejercido", f"${pres_ejercido:,.0f}"),
        ("Avance prom.", f"{avg_avance:.0f}%"),
    ]
    for i, (label, value) in enumerate(kpis):
        col = i + 1
        ds.cell(row=3, column=col, value=value).font = _kpi_value_font
        ds.cell(row=3, column=col).alignment = Alignment(horizontal="center")
        ds.cell(row=4, column=col, value=label).font = _kpi_label_font
        ds.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        ds.column_dimensions[get_column_letter(col)].width = 18

    # Estado distribution
    estado_counts: dict[str, int] = {}
    for o in obras:
        st = o.get("estado", "?")
        estado_counts[st] = estado_counts.get(st, 0) + 1

    ds["A7"] = "Por estado"
    ds["A7"].font = Font(bold=True, color=DARK, size=11)
    ds["A8"] = "Estado"
    ds["B8"] = "Cantidad"
    ds["A8"].font = _header_font
    ds["A8"].fill = _header_fill
    ds["B8"].font = _header_font
    ds["B8"].fill = _header_fill
    for i, (st, cnt) in enumerate(estado_counts.items()):
        ds.cell(row=9 + i, column=1, value=st)
        ds.cell(row=9 + i, column=2, value=cnt)

    if estado_counts:
        chart = PieChart()
        chart.title = "Estado de obras"
        chart.width = 14
        chart.height = 10
        end_row = 8 + len(estado_counts)
        chart.add_data(Reference(ds, min_col=2, min_row=8, max_row=end_row), titles_from_data=True)
        chart.set_categories(Reference(ds, min_col=1, min_row=9, max_row=end_row))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        ds.add_chart(chart, "D7")

    # Categoria distribution
    cat_counts: dict[str, int] = {}
    for o in obras:
        c = o.get("categoria_id", "?")
        cat_counts[c] = cat_counts.get(c, 0) + 1

    cat_start = 9 + len(estado_counts) + 3
    ds.cell(row=cat_start, column=1, value="Por categoría").font = Font(bold=True, color=DARK, size=11)
    ds.cell(row=cat_start + 1, column=1, value="Categoría").font = _header_font
    ds.cell(row=cat_start + 1, column=1).fill = _header_fill
    ds.cell(row=cat_start + 1, column=2, value="Cantidad").font = _header_font
    ds.cell(row=cat_start + 1, column=2).fill = _header_fill
    for i, (cat, cnt) in enumerate(cat_counts.items()):
        ds.cell(row=cat_start + 2 + i, column=1, value=cat)
        ds.cell(row=cat_start + 2 + i, column=2, value=cnt)

    if cat_counts:
        chart2 = BarChart()
        chart2.title = "Obras por categoría"
        chart2.width = 14
        chart2.height = 10
        end_row = cat_start + 1 + len(cat_counts)
        chart2.add_data(Reference(ds, min_col=2, min_row=cat_start + 1, max_row=end_row), titles_from_data=True)
        chart2.set_categories(Reference(ds, min_col=1, min_row=cat_start + 2, max_row=end_row))
        ds.add_chart(chart2, f"D{cat_start}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
